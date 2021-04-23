from selenium import webdriver
import time
from webdriver_manager.chrome import ChromeDriverManager  # permite simular uso do navegador sem baixar ferramentas
import openpyxl


# entrar no site da amazon
class Amazonscrapy:
    def __init__(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get('https://www.amazon.com.br/')
        self.aparelho = ['iphone']
        time.sleep(1)
        self.lista_nome_celulares = []
        self.lista_preco_celulares = []

    def iniciar(self):
        self.buscar_aparelho('Iphone')
        self.varrer_site()
        self.criar_planilha()

    def buscar_aparelho(self, aparelho):
        print('Buscando aparelho')
        self.driver.find_element_by_xpath('//*[@id="twotabsearchtextbox"]').send_keys(aparelho)
        self.driver.find_element_by_xpath('//*[@id="nav-search-submit-button"]').click()

    def varrer_site(self):
        print('Escaneando...')
        for i in range(0, 48):
            lista_nomes = self.driver.find_elements_by_xpath(
                '//span[@class="a-size-base-plus a-color-base a-text-normal"]')
            self.lista_nome_celulares.append(lista_nomes[i].text)
            time.sleep(0.2)
            lista_precos = self.driver.find_elements_by_xpath(
                '//span[@class="a-price-whole"]')
            self.lista_preco_celulares.append(lista_precos[i].text)
        print('Escaneamento completo')

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares.title = 'Celulares'
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preço'
        # zip funciona para trabalhar com multiplas listas
        for nome, preco in zip(self.lista_nome_celulares, self.lista_preco_celulares):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_nomes_valores.xlsx")
        self.driver.quit()
        print('Planilha criada')


escanear = Amazonscrapy()
escanear.iniciar()
